"""
Gera relatório financeiro em .xlsx com duas abas: lançamentos detalhados e resumo DRE.

Uso:
    python tools/export_financial_to_excel.py \
        --user_id <uuid> \
        --start_date 2026-05-01 \
        --end_date 2026-05-31

Saída:
    Arquivo .tmp/relatorio_2026-05-01_2026-05-31.xlsx
    Stdout JSON: {"file_path": ".tmp/relatorio_...", "entry_count": 47}

Dependências:
    pip install openpyxl asyncpg python-dotenv
"""

import argparse
import json
import os
from datetime import date
from pathlib import Path

import asyncpg
from dotenv import load_dotenv
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

load_dotenv()

HEADER_FILL = PatternFill("solid", fgColor="4A90D9")
HEADER_FONT = Font(bold=True, color="FFFFFF")
SUMMARY_LABEL_FONT = Font(bold=True)
PINK_FILL = PatternFill("solid", fgColor="F8D7E3")


def _style_header_row(ws, row: int, col_count: int):
    for col in range(1, col_count + 1):
        cell = ws.cell(row=row, column=col)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center")


def _br_currency(value: float) -> str:
    return f"R$ {value:,.2f}".replace(",", "X").replace(".", ",").replace("X", ".")


def _br_date(d: date) -> str:
    return d.strftime("%d/%m/%Y")


async def export_to_excel(user_id: str, start_date_str: str, end_date_str: str) -> dict:
    start_date = date.fromisoformat(start_date_str)
    end_date = date.fromisoformat(end_date_str)

    conn = await asyncpg.connect(os.environ["DATABASE_URL"])
    try:
        rows = await conn.fetch(
            """
            SELECT
                fe.entry_date,
                fe.client_name,
                fe.procedure_name,
                fe.revenue,
                fe.cost,
                fe.profit,
                a.is_paid,
                a.scheduled_at
            FROM financial_entries fe
            JOIN appointments a ON a.id = fe.appointment_id
            WHERE fe.user_id = $1
              AND fe.entry_date BETWEEN $2 AND $3
            ORDER BY a.scheduled_at ASC
            """,
            user_id, start_date, end_date,
        )

        total_revenue = sum(float(r["revenue"]) for r in rows)
        total_cost = sum(float(r["cost"]) for r in rows)
        total_profit = sum(float(r["profit"]) for r in rows)
        margin = (total_profit / total_revenue * 100) if total_revenue > 0 else 0.0

    finally:
        await conn.close()

    wb = Workbook()

    # --- Aba 1: Lançamentos ---
    ws1 = wb.active
    ws1.title = "Lançamentos"

    headers = ["Data", "Horário", "Cliente", "Procedimento", "Receita (R$)", "Custo (R$)", "Lucro (R$)", "Pago"]
    for col, header in enumerate(headers, start=1):
        ws1.cell(row=1, column=col, value=header)
    _style_header_row(ws1, 1, len(headers))

    for row_idx, r in enumerate(rows, start=2):
        ws1.cell(row=row_idx, column=1, value=_br_date(r["entry_date"]))
        ws1.cell(row=row_idx, column=2, value=r["scheduled_at"].strftime("%H:%M"))
        ws1.cell(row=row_idx, column=3, value=r["client_name"])
        ws1.cell(row=row_idx, column=4, value=r["procedure_name"])
        ws1.cell(row=row_idx, column=5, value=float(r["revenue"]))
        ws1.cell(row=row_idx, column=6, value=float(r["cost"]))
        ws1.cell(row=row_idx, column=7, value=float(r["profit"]))
        ws1.cell(row=row_idx, column=8, value="Sim" if r["is_paid"] else "Não")
        if row_idx % 2 == 0:
            for col in range(1, 9):
                ws1.cell(row=row_idx, column=col).fill = PatternFill("solid", fgColor="F2F2F2")

    # Ajustar largura das colunas
    col_widths = [12, 10, 25, 25, 14, 14, 14, 8]
    for col, width in enumerate(col_widths, start=1):
        ws1.column_dimensions[get_column_letter(col)].width = width

    # --- Aba 2: Resumo DRE ---
    ws2 = wb.create_sheet(title="Resumo DRE")

    ws2.column_dimensions["A"].width = 30
    ws2.column_dimensions["B"].width = 20

    ws2["A1"] = "DRE — Resumo do Período"
    ws2["A1"].font = Font(bold=True, size=14)
    ws2.merge_cells("A1:B1")

    summary_rows = [
        ("Período", f"{_br_date(start_date)} a {_br_date(end_date)}"),
        ("Atendimentos Realizados", len(rows)),
        ("", ""),
        ("(+) Receita Bruta", _br_currency(total_revenue)),
        ("(-) Custo Total", _br_currency(total_cost)),
        ("(=) Lucro Bruto", _br_currency(total_profit)),
        ("Margem de Lucro", f"{margin:.1f}%"),
    ]

    for row_offset, (label, value) in enumerate(summary_rows, start=3):
        cell_label = ws2.cell(row=row_offset, column=1, value=label)
        cell_value = ws2.cell(row=row_offset, column=2, value=value)
        cell_label.font = SUMMARY_LABEL_FONT
        if label == "(=) Lucro Bruto":
            cell_label.fill = PINK_FILL
            cell_value.fill = PINK_FILL

    Path(".tmp").mkdir(exist_ok=True)
    filename = f".tmp/relatorio_{start_date_str}_{end_date_str}.xlsx"
    wb.save(filename)

    return {"file_path": filename, "entry_count": len(rows)}


if __name__ == "__main__":
    import asyncio

    parser = argparse.ArgumentParser(description="Exporta relatório financeiro para Excel")
    parser.add_argument("--user_id", required=True)
    parser.add_argument("--start_date", required=True, help="YYYY-MM-DD")
    parser.add_argument("--end_date", required=True, help="YYYY-MM-DD")
    args = parser.parse_args()

    result = asyncio.run(export_to_excel(args.user_id, args.start_date, args.end_date))
    print(json.dumps(result, ensure_ascii=False, indent=2))
